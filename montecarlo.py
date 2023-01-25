# Trabajo realizado por
#  Jose Luis Rico Ramos y Miguel Tirado Guzmán,
# Basado en el trabajo de:
#   Cristina Aranda García-Torres
#   Paula Hernández Lozano

import numpy as np
import xlsxwriter
import pandas as pd
import math
from math import exp
import openpyxl
from SeminarioCovid.model import CovidModel



# Inicilizamos a vacio cada uno de los vectores con las variables exogenas, 
# que luego rellenaremos de datos en funcion de las simulaciones
n_ciudadanos = []
libre_movimiento = []
poblacion_contagios = []
capacidad_sanidad = []
tipo_cepa = []

# Trabajamos con un numero de 10 simulaciones
NUM_MUESTRAS = 10

# Damos valores aleatorios acada una de nuestras simulaciones, para poder ejecutar luego 
# nuestro modelo y comprobar cual de ellas se adapta mejor a los datos reales

for i in range (NUM_MUESTRAS):
    n_ciudadanos.append(np.random.randint(150,200))
    libre_movimiento.append(np.random.randint(1,100))
    poblacion_contagios.append(np.random.randint(1,100))
    capacidad_sanidad.append(np.random.randint(1,100))
    tipo_cepa.append(np.random.randint(1,3))

class Montecarlo():
    def __init__(self) -> None:
        self.Rvt_todos()
        #self.todos_montecarlo()


    #Esta función crea un archivo excel donde almacenará todos los datos que vaya creando para montecarlo
    def montecarlo(self, nombre_archivo):
        libro = xlsxwriter.Workbook(nombre_archivo)
        hoja = libro.add_worksheet()
        col =0
        row = 0
        hoja.write(row,col, 'n_ciudadanos'+str(col))
        # Iteramos los datos para ir pintando fila a fila
        for dato in (n_ciudadanos):
            hoja.write(row+1, col, dato)
            row += 1

        col += 1
        row = 0
        hoja.write(row,col, 'libre_movimiento'+str(col))
        # Iteramos los datos para ir pintando fila a fila
        for dato in (libre_movimiento):
            hoja.write(row+1, col, dato)
            row += 1

        col += 1
        row = 0
        hoja.write(row,col, 'poblacion_contagios'+str(col))
        # Iteramos los datos para ir pintando fila a fila
        for dato in (poblacion_contagios):
            hoja.write(row+1, col, dato)
            row += 1

        col += 1
        row = 0
        hoja.write(row,col, 'capacidad_sanidad'+str(col))
        # Iteramos los datos para ir pintando fila a fila
        for dato in (capacidad_sanidad):
            hoja.write(row+1, col, dato)
            row += 1

        col += 1
        row = 0
        hoja.write(row,col, 'tipo_cepa'+str(col))
        # Iteramos los datos para ir pintando fila a fila
        for dato in (tipo_cepa):
            hoja.write(row+1, col, dato)
            row += 1

        #Cerramos el libro
        libro.close()

    #Genera las 10 muestras necesarias para Montecarlo
    def todos_montecarlo(self):
        for i in range(10):
            self.montecarlo("Montecarlo_covid"+str(i)+".xlsx")

    # Calculamos con esta funcion el error cuadratico medio respecto a los datos reales
    def mse(self,vec1, vec2):
        return sum((np.array(vec1) - np.array(vec2))**2) / len(vec1)

    def Rvt_todos(self):
        #Inicializamos el error a un numero muy grande para que se actualice
        menor_error = 9999999999999999
        # Abrir el archivo XLSX
        wb = openpyxl.load_workbook('cifras_contagios.xlsx')

        # Seleccionar la hoja de trabajo que queremos leer
        ws = wb['Hoja1']

        # Crear una lista para almacenar los valores leídos
        contagios_reales = []

        # Iterar a través de las 40 primeras filas de la columna "España"
        for row in ws['C41:C80']:
            # Agregar el valor de la celda a la lista
            contagios_reales.append(row[0].value)
        # Cerrar el archivo XLSX
        wb.close()

        #CREACION DE UN EXCEL EN EL QUE HAY UNA HOJA POR CADA VEZ QUE EJECUTAMOS LA SIMULACION
        # De querer almacenar las iteraciones en un excel, lo primero que deberiamos hacer es crearnos 
        # el excel y abrirlo dentro del programa gracias a la libreria pandas
            
        libro = xlsxwriter.Workbook("ArchivoSimulacion.xlsx")
        col_ = 0
        hoja = libro.add_worksheet("Simulacion ")
        for i in range (NUM_MUESTRAS):
            
            modelo = CovidModel(n_ciudadanos[i],
                                        libre_movimiento[i],
                                        capacidad_sanidad[i],
                                        poblacion_contagios[i],
                                        tipo_cepa[i],20,20)
            contagios_simulacion=[]
            
            
            # el numero de steps lo dejamos a 40 por falta de tiempo, ya que en funcion del numero de row que 
            # queramos analizar el tiempo es de tipo O(n^2) cuadratico al tener
            
            
            # el bucle anidado, por lo que podría llegar a tardar mucho tiempo en finalizar.
                
            
            # Despues cada vez que quisieramos escribir en el excel, deberiamos de utilizar la funcion write dentro del for, variando la row y col que son las filas y las 
            # columnas del excel, en funcion de la iteracion en la que estemos y con la variable con la que estemos trabajando
            row_hoja=0
            hoja.write(row_hoja,col_, "Contagios Simulacion "+str(i+1))
            # hoja.write(row_hoja,1, "No contagiados")
            # hoja.write(row_hoja,2, "Fallecidos")
            row_hoja=1
            for j in range (40):
                
                #Así escribimos por fila el numero de contagiados, sanos y fallecidos por iteracion
                hoja.write(row_hoja,col_, str(modelo.current_non_healthy_agents(modelo) * (235000*math.log(j+1,8)/n_ciudadanos[i])))
                # hoja.write(row_hoja,1, str(modelo.current_healthy_agents(modelo)))
                # hoja.write(row_hoja,2, str(modelo.current_dead_agents(modelo))) 
                row_hoja+=1               
                modelo.step()
                contagios_simulacion.append(modelo.current_non_healthy_agents(modelo) *(235000*math.log(j+1,8)/n_ciudadanos[i])) 

            #Para esa simulación, una vez tenemos los contagios los comparamos con los reales
            # y calculamos su error cuadratico medio
            error_cuad = self.mse(contagios_simulacion, contagios_reales)

            #Si el error de esta simulacion es la menor, actualizamos cual es la simulacion con menor error
            if(menor_error > error_cuad):
                menor_error = error_cuad
                row_ = i
            #print(error_cuad , i)
            col_+=1
                    
        hoja.write(0,col_, "Contagios reales")

        row_hoja = 0
        for dato in (contagios_reales):
            hoja.write(row_hoja+1,col_, dato)
            row_hoja += 1
            
        libro.close()
        print ("Numero ciudadanos: " , n_ciudadanos[row_] ,"Porcentaje libre movimiento: " ,  libre_movimiento[row_], "Capacidad sanidad: " , capacidad_sanidad[row_], "Poblacion inicialmente contagiada: " ,poblacion_contagios[row_], "Tipo de cepa: ",tipo_cepa[row_], "Iteracion donde se consigue: ", row_)

ejecutar = Montecarlo()